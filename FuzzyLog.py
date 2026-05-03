import openpyxl
import sys


class FuzzyLog:

    def __init__(self):
        self.servis_params = {
            'sangat_buruk': (1, 1, 20, 35),      # (a, b, c, d) untuk trapezoid
            'buruk': (20, 35, 40, 50),
            'sedang': (40, 50, 60, 70),
            'baik': (60, 70, 80, 90),
            'sangat_baik': (80, 90, 100, 100)
        }
        self.harga_params = {
            'sangat_murah': (25000, 25000, 30000, 35000),
            'murah': (30000, 35000, 38000, 42000),
            'sedang': (38000, 42000, 46000, 50000),
            'mahal': (46000, 50000, 52000, 55000),
            'sangat_mahal': (52000, 55000, 55000, 55000)
        }
        self.kelayakan_params = {
            'sangat_tidak_layak': (0, 0, 15, 25),
            'tidak_layak': (15, 25, 35, 45),
            'cukup_layak': (35, 45, 55, 65),
            'layak': (55, 65, 75, 85),
            'sangat_layak': (75, 85, 100, 100)
        }
        self.rules = self._createRules()
    
    def _createRules(self):
        rules = {
            # Servis Sangat Buruk
            ('sangat_buruk', 'sangat_murah'): 'tidak_layak',
            ('sangat_buruk', 'murah'): 'sangat_tidak_layak',
            ('sangat_buruk', 'sedang'): 'sangat_tidak_layak',
            ('sangat_buruk', 'mahal'): 'sangat_tidak_layak',
            ('sangat_buruk', 'sangat_mahal'): 'sangat_tidak_layak',
            
            # Servis Buruk
            ('buruk', 'sangat_murah'): 'cukup_layak',
            ('buruk', 'murah'): 'tidak_layak',
            ('buruk', 'sedang'): 'tidak_layak',
            ('buruk', 'mahal'): 'sangat_tidak_layak',
            ('buruk', 'sangat_mahal'): 'sangat_tidak_layak',
            
            # Servis Sedang
            ('sedang', 'sangat_murah'): 'layak',
            ('sedang', 'murah'): 'cukup_layak',
            ('sedang', 'sedang'): 'cukup_layak',
            ('sedang', 'mahal'): 'tidak_layak',
            ('sedang', 'sangat_mahal'): 'tidak_layak',
            
            # Servis Baik
            ('baik', 'sangat_murah'): 'sangat_layak',
            ('baik', 'murah'): 'sangat_layak',
            ('baik', 'sedang'): 'layak',
            ('baik', 'mahal'): 'cukup_layak',
            ('baik', 'sangat_mahal'): 'tidak_layak',
            
            # Servis Sangat Baik
            ('sangat_baik', 'sangat_murah'): 'sangat_layak',
            ('sangat_baik', 'murah'): 'sangat_layak',
            ('sangat_baik', 'sedang'): 'sangat_layak',
            ('sangat_baik', 'mahal'): 'layak',
            ('sangat_baik', 'sangat_mahal'): 'cukup_layak',
        }
        return rules
    
    def trapezoid(self, x, a, b, c, d):
        if x <= a or x >= d:
            return 0.0
        elif a < x <= b:
            return (x - a) / (b - a) if b != a else 1.0
        elif b < x <= c:
            return 1.0
        elif c < x < d:
            return (d - x) / (d - c) if d != c else 1.0
        return 0.0
    
    def fuzzification(self, servis, harga):

        servis_fuzzy = {}
        for kategori, params in self.servis_params.items():
            servis_fuzzy[kategori] = self.trapezoid(servis, *params)
        
        # Fuzzifikasi Harga
        harga_fuzzy = {}
        for kategori, params in self.harga_params.items():
            harga_fuzzy[kategori] = self.trapezoid(harga, *params)
        
        return servis_fuzzy, harga_fuzzy
    
    def inference(self, servis_fuzzy, harga_fuzzy):
        output_fuzzy = {
            'sangat_tidak_layak': 0.0,
            'tidak_layak': 0.0,
            'cukup_layak': 0.0,
            'layak': 0.0,
            'sangat_layak': 0.0
        }
        
        # Terapkan setiap aturan
        for (servis_cat, harga_cat), output_cat in self.rules.items():
            # Hitung firing strength menggunakan MIN (T-Norm standar Mamdani)
            firing_strength = min(servis_fuzzy[servis_cat], harga_fuzzy[harga_cat])
            
            # Gunakan MAX untuk agregasi (S-Norm standar Mamdani)
            output_fuzzy[output_cat] = max(output_fuzzy[output_cat], firing_strength)
        
        return output_fuzzy
    
    def defuzz(self, output_fuzzy, servis, harga):

        resolution = 0.5
        x_values = []
        y_values = []
        
        x = 0.0
        while x <= 100.0:
            membership_at_x = 0.0
            
            for category, firing_strength in output_fuzzy.items():
                if firing_strength > 0:
                    a, b, c, d = self.kelayakan_params[category]
                    membership_trapezoid = self.trapezoid(x, a, b, c, d)
                    clipped = min(membership_trapezoid, firing_strength)
                    membership_at_x = max(membership_at_x, clipped)
            
            x_values.append(x)
            y_values.append(membership_at_x)
            x += resolution
        
        numerator = sum(x * y for x, y in zip(x_values, y_values))
        denominator = sum(y_values)
        
        if denominator == 0:
            base_score = 0.0
        else:
            base_score = numerator / denominator
        
        # Tambahkan faktor penyesuaian halus berdasarkan nilai input
        # Normalisasi servis (0-100) dan harga (25000-55000 -> 0-100 terbalik)
        servis_normalized = servis / 100.0  # 0-1
        harga_normalized = 1.0 - ((harga - 25000) / 30000)  # 1(murah)-0(mahal)
        
        # Faktor penyesuaian kecil (maksimal ±5 poin)
        adjustment = (servis_normalized * 2.5) + (harga_normalized * 2.5)
        
        final_score = base_score + adjustment
        
        # Pastikan dalam rentang 0-100
        return max(0.0, min(100.0, final_score))
    
    def read(self, filename):

        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            
            data = []
            # Baca dari baris ke-2 (baris 1 adalah header)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # check ID
                    data.append({
                        'id': int(row[0]),
                        'kualitas_servis': float(row[1]),
                        'harga': float(row[2])
                    })
            
            workbook.close()
            return data
            
        except FileNotFoundError:
            print(f"✗ Error: File {filename} tidak ditemukan!")
            sys.exit(1)
        except Exception as e:
            print(f"✗ Error saat membaca file: {str(e)}")
            sys.exit(1)
    
    def processRes(self, data):

        results = []
        
        for resto in data:
            # 1. Fuzzification
            servis_fuzzy, harga_fuzzy = self.fuzzification(
                resto['kualitas_servis'], 
                resto['harga']
            )
            
            # 2. Inferensi
            output_fuzzy = self.inference(servis_fuzzy, harga_fuzzy)
            
            # 3. Defuzzification
            skor = self.defuzz(output_fuzzy, resto['kualitas_servis'], resto['harga'])
            
            # Simpan hasil
            results.append({
                'id': resto['id'],
                'kualitas_servis': resto['kualitas_servis'],
                'harga': resto['harga'],
                'skor_kelayakan': skor
            })
        
        return results
    
    def select5(self, results):

        sorted_results = sorted(results, key=lambda x: x['skor_kelayakan'], reverse=True)
        
        # Ambil 5 teratas
        top_5 = sorted_results[:5]
        
        return top_5
    
    def save(self, top_5, filename):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Top 5 Restoran"

            headers = ['ID Restoran', 'Kualitas Servis', 'Harga', 'Skor Kelayakan']
            sheet.append(headers)

            for cell in sheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)

            for resto in top_5:
                sheet.append([
                    resto['id'],
                    resto['kualitas_servis'],
                    resto['harga'],
                    round(resto['skor_kelayakan'], 2)
                ])

            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 18
            sheet.column_dimensions['C'].width = 12
            sheet.column_dimensions['D'].width = 18

            workbook.save(filename)
            workbook.close()
            
            
        except Exception as e:
            print(f"✗ Error saat menyimpan file: {str(e)}")
            sys.exit(1)
    
    def display(self, top_5):

        print(f"{'Rank':<6} {'ID':<8} {'Kualitas Servis':<18} {'Harga':<12} {'Skor':<10}")
        print("-"*70)
        
        for i, resto in enumerate(top_5, 1):
            print(f"{i:<6} {resto['id']:<8} {resto['kualitas_servis']:<18.1f} "
                  f"{resto['harga']:<12,.0f} {resto['skor_kelayakan']:<10.2f}")
        


def main():
    input_file = "restoran.xlsx"
    output_file = "peringkat.xlsx"

    fuzzy_system = FuzzyLog()
    data = fuzzy_system.read(input_file)
    results = fuzzy_system.processRes(data)
    top_5 = fuzzy_system.select5(results)
    fuzzy_system.display(top_5)
    fuzzy_system.save(top_5, output_file)
    
    print("test debug")


if __name__ == "__main__":
    main()
